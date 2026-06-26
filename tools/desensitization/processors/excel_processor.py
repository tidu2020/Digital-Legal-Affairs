"""
Excel 处理器（.xlsx）
基于 openpyxl，遍历所有工作表的单元格，
对单元格值做替换，保留单元格格式（字体、边框、填充、公式等）。
"""
import logging
from typing import List

from openpyxl import load_workbook

from core.llm_client import Replacement
from processors.base import BaseProcessor

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    """Excel .xlsx 处理器"""

    def extract_text(self, file_path: str) -> str:
        # data_only=True 以读取公式计算后的值
        wb = load_workbook(file_path, data_only=True)
        parts: List[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        parts.append(str(cell.value))
        wb.close()
        return "\n".join(parts)

    def apply_replacements(
        self, file_path: str, replacements: List[Replacement]
    ) -> List[Replacement]:
        # 保留格式：不使用 data_only，保留公式
        wb = load_workbook(file_path)
        applied_set = set()
        applied: List[Replacement] = []

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    # 仅处理字符串类型单元格（数字/公式不做文本替换）
                    if not isinstance(cell.value, str):
                        continue
                    new_text, used = self._apply_text_replacements(
                        cell.value, replacements
                    )
                    if new_text != cell.value:
                        cell.value = new_text
                        for r in used:
                            self._track_applied(applied_set, applied, r)

        wb.save(file_path)
        wb.close()
        return applied
