"""
共享文件处理工具 - 支持多格式文件解析和信息提取
"""
import io
import os
from typing import List, Dict, Any, Optional


def extract_text_from_file(file_content: bytes, filename: str) -> str:
    """从文件内容提取文本"""
    ext = os.path.splitext(filename)[1].lower()
    
    if ext in ('.txt', '.md'):
        return file_content.decode("utf-8", errors="ignore")
    
    elif ext == '.docx':
        try:
            from docx import Document
            doc = Document(io.BytesIO(file_content))
            return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        except Exception as e:
            return f"[DOCX解析失败: {e}]"
    
    elif ext == '.pdf':
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(file_content))
            texts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    texts.append(text)
            return "\n".join(texts)
        except Exception as e:
            return f"[PDF解析失败: {e}]"
    
    elif ext in ('.xls', '.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            texts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        texts.append(row_text)
            return "\n".join(texts)
        except Exception as e:
            return f"[Excel解析失败: {e}]"
    
    elif ext in ('.jpg', '.jpeg', '.png', '.bmp', '.tiff'):
        return f"[图片文件: {filename} - 需要OCR处理]"
    
    else:
        return f"[不支持的文件格式: {ext}]"


def extract_info_with_llm(llm_client, text: str, file_type: str = "document") -> Dict[str, Any]:
    """使用LLM从文本中提取结构化信息"""
    system_prompt = """你是一个信息提取专家。请从给定的文档内容中提取关键信息。

输出JSON格式：
{
  "title": "文档标题",
  "summary": "文档摘要（100字以内）",
  "key_points": ["关键点1", "关键点2", ...],
  "dates": ["提到的日期"],
  "parties": ["涉及的各方"],
  "amounts": ["涉及的金额"],
  "laws": ["引用的法律法规"],
  "risks": ["识别的风险点"],
  "action_items": ["需要执行的事项"]
}

只输出纯JSON，不要其他文字。"""
    
    try:
        truncated_text = text[:8000] if len(text) > 8000 else text
        result = llm_client.chat_json(system_prompt, f"文档类型：{file_type}\n\n文档内容：\n{truncated_text}")
        return result
    except Exception as e:
        return {"error": str(e), "summary": text[:200]}


def process_files(llm_client, files: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    批量处理多个文件
    
    Args:
        files: [{"filename": str, "content": bytes}, ...]
    
    Returns:
        {
            "files": [{"filename": str, "text": str, "info": dict}, ...],
            "combined_text": str,
            "combined_info": dict
        }
    """
    processed_files = []
    all_texts = []
    
    for file_info in files:
        filename = file_info["filename"]
        content = file_info["content"]
        
        text = extract_text_from_file(content, filename)
        all_texts.append(f"=== {filename} ===\n{text}")
        
        processed_files.append({
            "filename": filename,
            "text": text,
            "size": len(content)
        })
    
    combined_text = "\n\n".join(all_texts)
    
    combined_info = {}
    if llm_client and combined_text.strip():
        combined_info = extract_info_with_llm(llm_client, combined_text)
    
    return {
        "files": processed_files,
        "combined_text": combined_text,
        "combined_info": combined_info,
        "file_count": len(files),
        "total_chars": len(combined_text)
    }
